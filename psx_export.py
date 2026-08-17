#!/usr/bin/env python3
"""
psx_export.py — Excel and PDF export of the analysis.

NEWLY WRITTEN MODULE. The original was not supplied with the v2.0 bundle.
Reimplemented against build_excel(symbols, bench, partial) and
build_pdf(symbols, bench, partial), both returning an in-memory buffer.

openpyxl and reportlab are imported lazily and their ImportError is allowed
to propagate — the web layer catches it and tells the user which package to
install, which is more useful than a generic failure.
"""

import io
from datetime import datetime

import psx_brain
import psx_report

DISCLAIMER = ("Decision support only. Every level here is computed from price "
              "and volume by unvalidated thresholds — confirm manually before "
              "placing any order.")

COLUMNS = [
    ("Symbol", "symbol"), ("Price", "price"), ("Verdict", "verdict"),
    ("Score", "score"), ("Confidence", "confidence"), ("Daily trend", "_dTrend"),
    ("Weekly trend", "_wTrend"), ("Cloud", "_cloud"), ("RSI", "_rsi"),
    ("ADX", "_adx"), ("Daily flow", "_dVol"), ("Weekly flow", "_wVol"),
    ("Trigger", "_trigger"), ("Stop", "_stop"), ("Target 1", "_t1"),
    ("Target 2", "_t2"), ("Risk %", "_risk_pct"), ("As of", "asof"),
]


def _analyse_all(symbols, bench=None, partial="drop"):
    """Analyse each symbol, keeping failures as an error row rather than
    dropping them — an export that silently omits a stock is worse than one
    that says it could not read it."""
    out = []
    for s in symbols:
        sym = str(s).upper()
        try:
            df = psx_report.load_from_psx(sym, 3)
            out.append(psx_brain.analyse(sym, df, bench, partial))
        except Exception as e:
            out.append({"symbol": sym, "error": f"{type(e).__name__}: {e}"[:160]})
    return out


def _flat(res):
    """Flatten one result into the column keys above."""
    if "error" in res:
        return {"symbol": res["symbol"], "verdict": "NO DATA",
                "price": None, "score": None, "confidence": None,
                "asof": res["error"]}
    st, lv = res["state"], res["levels"]
    row = {"symbol": res["symbol"], "price": res["price"],
           "verdict": res["verdict"], "score": res["score"],
           "confidence": res["confidence"], "asof": res.get("asof")}
    row.update({f"_{k}": st[k] for k in
                ("dTrend", "wTrend", "cloud", "rsi", "adx", "dVol", "wVol")})
    row.update({f"_{k}": lv[k] for k in
                ("trigger", "stop", "t1", "t2", "risk_pct")})
    return row


def build_excel(symbols, bench=None, partial="drop"):
    """Workbook: one summary sheet plus a written read per symbol."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    results = _analyse_all(symbols, bench, partial)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3864")
    for i, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font, cell.fill = head, fill
        cell.alignment = Alignment(horizontal="center")

    tone = {"BUY": "C6EFCE", "BUY ON TRIGGER": "FFEB9C",
            "WAIT": "F2F2F2", "AVOID": "FFC7CE", "NO DATA": "D9D9D9"}
    for r, res in enumerate(results, start=2):
        row = _flat(res)
        for i, (_, key) in enumerate(COLUMNS, start=1):
            ws.cell(row=r, column=i, value=row.get(key))
        colour = tone.get(row.get("verdict"))
        if colour:
            ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=colour)

    for i, (label, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(11, len(label) + 3)
    ws.freeze_panes = "A2"

    note = ws.cell(row=len(results) + 3, column=1, value=DISCLAIMER)
    note.font = Font(italic=True, size=9)

    for res in results:
        if "error" in res:
            continue
        name = res["symbol"][:31]
        sh = wb.create_sheet(name)
        sh.column_dimensions["A"].width = 110
        lines = [f"{res['symbol']} — {res['verdict']} "
                 f"(score {res['score']:+.1f}, confidence {res['confidence']}/100)",
                 "", res["summary"], ""]
        for title, items in (("SUPPORTING THE TRADE", res["bull"]),
                             ("AGAINST THE TRADE", res["bear"]),
                             ("WATCH OUT", res["flags"])):
            if items:
                lines.append(title)
                lines += [f"  - {x}" for x in items]
                lines.append("")
        lines += ["", DISCLAIMER]
        for i, line in enumerate(lines, start=1):
            c = sh.cell(row=i, column=1, value=line)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if i == 1:
                c.font = Font(bold=True, size=13)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(symbols, bench=None, partial="drop"):
    """A4 report: summary table, then one section per symbol."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (PageBreak, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    results = _analyse_all(symbols, bench, partial)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm,
                            title="PSX Research Terminal")
    ss = getSampleStyleSheet()
    body = ParagraphStyle("body", parent=ss["Normal"], fontSize=9,
                          leading=13, alignment=TA_LEFT)
    small = ParagraphStyle("small", parent=body, fontSize=7.5,
                           textColor=colors.grey)

    story = [Paragraph("PSX Research Terminal", ss["Title"]),
             Paragraph(f"{len(results)} stock(s) · generated "
                       f"{datetime.now():%d %b %Y %H:%M}", small),
             Spacer(1, 6 * mm)]

    cols = ["Symbol", "Price", "Verdict", "Score", "Conf", "Trigger", "Stop", "T1"]
    data = [cols]
    for res in results:
        row = _flat(res)
        data.append([row.get("symbol"), row.get("price"), row.get("verdict"),
                     row.get("score"), row.get("confidence"),
                     row.get("_trigger"), row.get("_stop"), row.get("_t1")])
    table = Table(data, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    story += [table, Spacer(1, 4 * mm), Paragraph(DISCLAIMER, small)]

    for res in results:
        if "error" in res:
            continue
        story.append(PageBreak())
        story.append(Paragraph(f"{res['symbol']} — {res['verdict']}", ss["Heading1"]))
        story.append(Paragraph(
            f"Score {res['score']:+.1f} · confidence {res['confidence']}/100 · "
            f"as of {res.get('asof')}", small))
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(res["summary"], body))
        for title, items in (("Supporting the trade", res["bull"]),
                             ("Against the trade", res["bear"]),
                             ("Watch out", res["flags"])):
            if not items:
                continue
            story += [Spacer(1, 3 * mm), Paragraph(f"<b>{title}</b>", body)]
            story += [Paragraph(f"• {x}", body) for x in items]
        story += [Spacer(1, 4 * mm), Paragraph(DISCLAIMER, small)]

    doc.build(story)
    buf.seek(0)
    return buf
